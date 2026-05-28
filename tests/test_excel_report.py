from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from src.reports.excel_report import generate_excel_report


def test_excel_report_has_required_sheets_and_safe_external_data(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("src.reports.excel_report.REPORTS_DIR", tmp_path)

    telemetry = pd.DataFrame(
        [
            {
                "timestamp": "2026-05-27T00:00:00",
                "mission_elapsed_min": 0,
                "phase": "PRE_LAUNCH",
                "battery_percent": 99,
                "oxygen_percent": 99,
                "fuel_percent": 99,
                "cabin_pressure_kpa": 101,
                "cabin_temperature_c": 22,
                "radiation_msv": 0.2,
                "communication_latency_ms": 500,
                "signal_strength_percent": 95,
                "engine_temperature_c": 600,
                "velocity_m_s": 0,
                "altitude_km": 0,
            }
        ]
    )
    alerts = pd.DataFrame(columns=["timestamp", "metric", "severity", "message", "value"])
    summary = {
        "run_id": "run-123",
        "generated_at": "2026-05-27T00:00:00",
        "mission_name": "ARTEMIS SENTINEL",
        "risk_score": 10,
    }
    external = {
        "power_data": {
            "source": "NASA_POWER_FALLBACK",
            "latitude": 28.5721,
            "longitude": -80.6480,
            "records_count": 7,
            "missing_records": 0,
            "missing_ratio": 0.0,
            "fallback_used": True,
            "fallback_reason": "TEST",
        },
        "donki_data": {
            "source": "NASA_DONKI_FALLBACK",
            "event_count": 1,
            "max_class_type": "C2.1",
            "fallback_used": True,
            "fallback_reason": "TEST",
            "events": [
                {
                    "event_id": "E1",
                    "begin_time": "t",
                    "peak_time": "t",
                    "class_type": "C2.1",
                    "source_location": "N00",
                    "active_region": "AR1",
                }
            ],
        },
    }

    report_path = generate_excel_report(telemetry, alerts, summary, external)
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    assert set(["mission_summary", "telemetry", "alerts", "external_data"]).issubset(wb.sheetnames)

    ws = wb["external_data"]
    content = []
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        content.extend([str(cell) for cell in row if cell is not None])
    text = " ".join(content)
    assert "api_key" not in text.lower()
    assert "-999" not in text

